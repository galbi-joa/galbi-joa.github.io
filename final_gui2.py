"""
Copyright 2024.07.28 -- 정용호
"""
import sys
import openpyxl
import datetime
import os
from shutil import copyfile
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QLabel,
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QProgressBar,
    QTextEdit,
    QMessageBox,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont

# --- 계정 매핑 및 계산 정보 ---
account_sums_bs = {
    'BS0069': ['미수금', '공사미수금', '분양미수금'], 
    'BS0070': ['단기대여금', '가지급금'],
    'BS0073': ['선급공사원가', '선급비용'],
    'BS0017': ['기타(당좌자산)','매도가능증권'],
    # ---
    'BS0076': ['원재료', '원자재'],
    'BS0078': ['제품', '완성주택'],
    'BS0080': ['반제품', '미완성주택'],
    'BS0082': ['상품', '미성공사', '기타(재고자산)', '저장품', '용지'],
    # ---
    'BS0088': ['기타(투자자산)'],
    # ---
    'BS0093': ['기계장치', '시설장치', '금형', '비품','공구와기구'],
    'BS0096': ['기타(유형자산)'],
    # ---
    'BS0105': ['이연법인세차'],
    'BS0107': ['기타(비유동자산)','기타(기타비유동자산)'],
    # ---
    'BS0111': ['유동성장기부채'],
    'BS0121': ['선수금', '공사선수금', '분양선수금'],
    # ---
    'BS0048': ['예수금', '가수금', '유동성금융리스부채', '기타(유동부채)'],
    # ---
    'BS0129': ['사채', '사채할인발행자금'],
    # ---
    'BS0136': ['퇴직급여충당금'],
    'BS0137': ['임대보증금', '이연법인세대', '기타(비유동부채)', '기타(고정부채)','기타부채성충당금1'],
    # ---
    'BS0140': ['주식발행초과금', '감자차익', '기타자본잉여금'],
    'BS0141': ['자본조정', "기타포괄손익누계액"],
    
    'YH0001': ['해와사업환산이익', '해외사업환산손실', '투자증권평가이익', '투자증권평가손실'],
    # --- 1:1 대응
    'BS0003': ['현금및현금성자산', '현금 및 현금성자산'],
    'BS0065': ['사용제한된 현금'],
    'BS0066': ['단기금융상품'],
    'BS0068': ['단기투자증권'],
    'BS0006': ['매출채권'],
    'BS0005': ['받을어음'],
    'BS0072': ['미수수익'],
    'BS0009': ['선급금'],
    'BS0077': ['재공품'],
    'BS0079': ['미착품'],
    'BS0081': ['재고충당금'],
    'BS0018': ['장기금융상품'],
    'BS0085': ['장기투자증권'],
    'BS0086': ['장기대여금'],
    'BS0087': ['투자부동산'],
    'BS0090': ['토지'],
    'BS0091': ['건물'],
    'BS0092': ['구축물'],
    'BS0094': ['차량운반구'],
    'BS0095': ['건설중인자산'],
    'BS0098': ['영업권'],
    'BS0183': ['개발비'],
    'BS0099': ['창업비'],
    'BS0101': ['산업재산권','기타(무형자산)'],
    'BS0103': ['장기재고자산'],
    'BS0104': ['보증금'],
    'BS0106': ['장기성매출채권'],
    'BS0109': ['단기차입금'], 
    'BS0110': ['관계사 차입금'],
    'BS0113': ['유동성사채'],
    'BS0114': ['전환사채'],
    'BS0115': ['기타 무보증사채'],
    'BS0116': ['기타 담보부사채'],
    'BS0118': ['매입채무'],
    'BS0119': ['미지급금'],
    'BS0120': ['미지급비용'],
    'BS0122': ['미지급법인세'],
    'BS0123': ['선수수익'],
    'BS0124': ['이연수익'],
    'BS0125': ['부채성 충당금', '부채성충당금'],
    'BS0134': ['후순위채'],
    'BS0133': ['금융리스부채'],
    'BS0135': ['장기성매입채무'],
    'BS0139': ['납입자본금=자본금'],
    'BS0167': ['자본준비금'],
    'BS0143': ['해외사업환산이익','투자증권평가이익'], 
    "BS0002": ["유동자산"],
    "BS0126": ['장기차입금'],
    "BS0142": ['재평가적립금'],
    "BS0059": ['이익잉여금']
}


account_sums_ls = {
    "PL0001": ["상품매출액", '제품매출액', '공사매출액', '분양매출액', '기타매출액'],
    "PL0002": ["매출원가"],
    "PL0011": ['상품매출원가', '제품매출원가', '공사원가', '분양원가', '기타원가'],
    "PL0015": ["매출총이익"],
    "PL0118": ["판매비와관리비"],
    "PL0038": ["급료와임금"],
    "PL0039": ["퇴직급여"],
    "PL0040": ["복리후생비", "북리후생비"],
    "PL0041": ["도서인쇄비"],
    "PL0042": ["여비교통비"],
    "PL0043": ["통신비"],
    "PL0044": ["차량유지비"],
    "PL0045": ["경상개발비"],
    "PL0046": ["세금과공과"],
    "PL0047": ["보험료"],
    "PL0048": ["지급수수료"],
    "PL0049": ["소모품비"],
    "PL0050": ["대손상각비"],
    "PL0051": ["접대비"],
    "PL0052": ["감가상각비"],
    "PL0053": ["무형자산상각비"],
    "PL0054": ["수선비"],
    "PL0056": ["광고선전비"],
    "PL0055": ["미분양주택관리비", "수도광열비", "운반비", "기타(판관비)", "임차료"],
    "PL0009": ["영업이익"],
    "PL0012": ["영업외수익"],
    "PL0057": ["이자수익"],
    "PL0058": ["배당금수익"],
    "PL0059": ["임대료"],
    "PL0060": ["투자증권처분이익", "매도가능증권처분이익"], #투자증권처분이익" -> "매도가능증권처분이익"
    "PL0061": ["투자증권평가이익", "매도가능증권평가이익"], #"투자증권평가이익" -> "매도가능증권평가이익"
    "PL0062": ["외화환산이익"],
    "PL0063": ["매도가능증권감액손실환입"],
    "PL0064": ["투자자산처분이익"],
    "PL0065": ["유형자산처분이익"],
    "PL0066": ["사채상환이익"],
    "PL0067": ["자산수증이익"],
    "PL0068": ["채무면제이익"],
    "PL0069": ["지분법평가이익"],
    "PL0070": ["외환차익", "보험차익", "법인세환급액", "대손충당금환입", "기타(영업외수익)"],
    "PL0013": ["영업외비용"],
    "PL0071": ["이자비용"],
    "PL0072": ["기타의대손상각비"],
    "PL0073": ["투자증권처분손실", "유가증권처분손실"],
    "PL0074": ["투자증권평가손실", "유가증권평가손실"],
    "PL0075": ["재고자산평가손실"],
    "PL0076": ["외화환산손실"],
    "PL0077": ["매도가능증권감액손실","투자유가증권감액손실"],
    "PL0078": ["투자자산처분손실"],
    "PL0079": ["유형자산처분손실"],
    "PL0080": ["사채상환손실"],
    "PL0081": ["지분법평가손실"],
    "PL0082": ["보상비"],
    "PL0083": ["법인세추납액"],
    "PL0084": ["외환차손", "기부금", "매출채권처분손실", "재해손실",'기타(영업외비용)'],
    "PL0004": ["8. 법인세차감전이익"],
    "PL0085": ["법인세비용"],
    "PL0086": ["10. 계속사업이익"],
    "PL0087": ["11. 중단사업손익"],
    "PL0016": ["법인세효과"],
    "PL0018": ["12. 당기순이익"],
    
}

# --- 함수 정의 ---
def sort_account_codes(account_sums):
    """
    account_sums 딕셔너리의 의존성 관계를 분석하여 계산 순서를 정렬합니다.
    순환 참조를 감지하고 처리합니다.
    """
    sorted_codes = []
    visited = set()
    processing = set()  # 현재 처리 중인 코드를 추적하는 집합

    def visit(code):
        if code in processing:
            raise ValueError(f"순환 참조 감지: {code}")
        if code in visited:
            return

        visited.add(code)
        processing.add(code)

        if code in account_sums:
            for source_code in account_sums[code]:
                visit(source_code)

        processing.remove(code)
        sorted_codes.append(code)

    for code in account_sums:
        try:
            visit(code)
        except ValueError as e:
            print(f"오류: {e}")  # 순환 참조 오류 출력
            # 순환 참조가 발생한 경우, 해당 코드를 sorted_codes에 추가하지 않고 건너뜁니다.

    return sorted_codes

def load_skd_data_for_company(sheet, row_start, row_end):
    """SKD 데이터를 불러와서 연도별, 헤더별로 값을 저장합니다.
    (account_mapping을 사용하지 않고 SKD 헤더 그대로 저장)
    """
    header_row = [cell.value.strip() if cell.value is not None else None for cell in sheet[1]]
    company_data = {}  # {year: {header: value, ...}, ...}

    for row_num in range(row_start, row_end + 1):
        data_row = [cell.value for cell in sheet[row_num]]
        year = str(data_row[1])[:4]  # "2020년" -> "2020"

        if year.isdigit():
            year = int(year)
            company_data[year] = {}
            for col_idx in range(6, len(data_row)):
                # 6번째 열 (인덱스 5)은 날짜/시간 데이터이므로 건너뜁니다.
                if col_idx == 5:
                    continue
                header_value = header_row[col_idx]
                # 헤더 그대로 저장
                company_data[year][header_value] = data_row[col_idx]
    return company_data

def calculate_sums(mapped_data, account_sums):
    """account_sums에 따라 값을 누적 계산합니다.
    SKD 파일의 계정 항목들을 기반으로 DNA 파일의 계정 항목을 계산합니다.
    """
    sorted_account_codes = sort_account_codes(account_sums)
    calculated_sums = {}

    for account_code in sorted_account_codes:
        if account_code in account_sums:
            calculated_sums[account_code] = 0
            for source_code in account_sums[account_code]:
                if source_code in mapped_data:
                    calculated_sums[account_code] += mapped_data[source_code]
                else:
                    print(f"Warning: SKD 데이터에서 '{source_code}' 항목을 찾을 수 없습니다.")

    # 빼야할 항목 모음.
    """
    if 'BS0059' in calculated_sums and '당기순이익-대차' in mapped_data:
        calculated_sums['BS0059'] -= mapped_data['당기순이익-대차']
    """
    if 'BS0143' in calculated_sums and '투자증권평가손실' in mapped_data:
        calculated_sums['BS0143'] -= mapped_data['투자증권평가손실']
    if 'BS0143' in calculated_sums and '해외사업환산손실' in mapped_data:
        calculated_sums['BS0143'] -= mapped_data['해외사업환산손실']
    
    return calculated_sums



def calculate_sums_for_company(company_data, account_sums):
    """각 연도별로 account_sums을 기반으로 누적 계산을 수행합니다."""
    calculated_sums = {}
    for year, year_data in company_data.items():
        calculated_sums[year] = calculate_sums(year_data, account_sums)
    return calculated_sums

def save_to_dna(calculated_sums, company_code, year, file_type, template_filepath, output_dir, log_output):
    """계산된 값을 DNA 파일에 저장합니다. 수식이 있는 셀은 건너뜁니다.
    파일명은 사업자번호_dna_[bs|is].xlsx 형식입니다.
    """
    output_filename = f"{company_code}_dna_{file_type}.xlsx"
    output_filepath = os.path.join(output_dir, output_filename)
    
    try:
        copyfile(template_filepath, output_filepath)
    except FileNotFoundError:
        log_output.append(f"Error: 기본 DNA {file_type.upper()} 파일을 찾을 수 없습니다. 파일 경로를 확인하세요.")
        return

    dna_workbook = openpyxl.load_workbook(output_filepath, data_only=False)
    dna_sheet = dna_workbook.active

    for year, year_data in calculated_sums.items():
        for row_idx in range(2, dna_sheet.max_row + 1):
            account_code = dna_sheet.cell(row=row_idx, column=2).value
            if account_code in year_data:
                for col_idx in range(3, 7):
                    year_cell_value = dna_sheet.cell(row=1, column=col_idx).value
                    if isinstance(year_cell_value, datetime.datetime):
                        year_cell_value = year_cell_value.strftime('%Y.%m.%d')
                    if year_cell_value is not None and year_cell_value.startswith(str(year) + "."):
                        if dna_sheet.cell(row=row_idx, column=col_idx).data_type != 'f':
                            dna_sheet.cell(row=row_idx, column=col_idx).value = year_data[account_code]
                        else:
                            log_output.append(f"Warning: {year}년 {account_code} 셀은 수식이 있어 값을 덮어쓰지 않습니다.")
                        break

    dna_workbook.save(output_filepath)
    log_output.append(f"--- 사업자 번호 {company_code} {file_type.upper()} 처리 완료 ---")


# --- 쓰레드 작업 클래스 ---
class WorkerThread(QThread):
    progress_updated = pyqtSignal(int)
    log_updated = pyqtSignal(str)
    finished = pyqtSignal()

    def __init__(self, skd_filepath, dna_bs_filepath, dna_is_filepath, output_dir):
        super().__init__()
        self.skd_filepath = skd_filepath
        self.dna_bs_filepath = dna_bs_filepath
        self.dna_is_filepath = dna_is_filepath
        self.output_dir = output_dir
        self.log_output = []

    def run(self):
        try:
            # skd.xlsx 파일 열기
            skd_workbook = openpyxl.load_workbook(self.skd_filepath, data_only=True)
            skd_sheet = skd_workbook.active

            # 사업자 번호와 결산 기준일 추출
            company_data_ranges = {}
            current_company_code = None
            start_row = 2
            for row_idx in range(2, skd_sheet.max_row + 1):
                company_code = skd_sheet.cell(row=row_idx, column=1).value

                # try-except 블록을 사용하여 ValueError 처리
                try:
                    year = int(str(skd_sheet.cell(row=row_idx, column=2).value)[:4])
                except ValueError:
                    self.log_output.append(f"Warning: 행 {row_idx}의 결산 기준일이 잘못되었습니다. 해당 행을 건너뜁니다.")
                    continue  # 현재 행 건너뛰기

                if current_company_code is None:
                    current_company_code = company_code

                if company_code != current_company_code:
                    company_data_ranges[current_company_code] = (start_row, row_idx - 1)
                    current_company_code = company_code
                    start_row = row_idx

            # 마지막 회사 정보 추가
            company_data_ranges[current_company_code] = (start_row, skd_sheet.max_row)

            total_companies = len(company_data_ranges)
            processed_companies = 0

            # 각 사업자별로 데이터 처리 및 DNA 파일 생성
            for company_code, (start_row, end_row) in company_data_ranges.items():
                self.log_output.append(f"--- 사업자 번호 {company_code} 처리 시작 ---")

                # BS 데이터 처리
                company_data_bs = load_skd_data_for_company(skd_sheet, start_row, end_row)
                calculated_sums_bs = calculate_sums_for_company(company_data_bs, account_sums_bs)
                year = int(str(skd_sheet.cell(row=start_row, column=2).value)[:4])
                save_to_dna(calculated_sums_bs, company_code, year, "bs", self.dna_bs_filepath, self.output_dir, self.log_output)

                # LS 데이터 처리
                company_data_ls = load_skd_data_for_company(skd_sheet, start_row, end_row)
                calculated_sums_ls = calculate_sums_for_company(company_data_ls, account_sums_ls)
                save_to_dna(calculated_sums_ls, company_code, year, "is", self.dna_is_filepath, self.output_dir, self.log_output)

                processed_companies += 1
                progress = int((processed_companies / total_companies) * 100)
                self.progress_updated.emit(progress)
                self.log_output.append(f"--- 사업자 번호 {company_code} 처리 완료 ---")

            for log_line in self.log_output:
                self.log_updated.emit(log_line)

            skd_workbook.close()

        except Exception as e:
            self.log_updated.emit(f"Error: {e}")

        finally:
            self.finished.emit()

# --- 메인 윈도우 클래스 ---
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("노가다 헬퍼 Ver 0.5")
        self.setGeometry(100, 100, 800, 600)  # 윈도우 크기 확장

        # 레이아웃 설정
        main_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        # --- 왼쪽 레이아웃 ---
        # 토끼 아스키 코드 (크기 조정)
        self.rabbit_label = QLabel(
            "      (\_/) \n"
            "      (='.'=) \n"
            "      (\")_(\")"
        )
        self.rabbit_label.setAlignment(Qt.AlignCenter)
        font = QFont()
        font.setPointSize(16)
        self.rabbit_label.setFont(font)

        # 파일 선택 버튼
        self.skd_button = QPushButton("Data 파일 선택")
        self.skd_button.clicked.connect(self.select_skd_file)
        self.dna_bs_button = QPushButton("DNA BS 파일 선택")
        self.dna_bs_button.clicked.connect(self.select_dna_bs_file)
        self.dna_is_button = QPushButton("DNA IS 파일 선택")
        self.dna_is_button.clicked.connect(self.select_dna_is_file)
        self.output_dir_button = QPushButton("결과 파일 저장 위치 선택")
        self.output_dir_button.clicked.connect(self.select_output_dir)

        # 파일 경로 표시 라벨
        self.skd_label = QLabel("Data 파일: ")
        self.dna_bs_label = QLabel("DNA BS 파일: ")
        self.dna_is_label = QLabel("DNA IS 파일: ")
        self.output_dir_label = QLabel("저장 위치: ")

        # 작업 시작 버튼
        self.start_button = QPushButton("작업 시작!")
        self.start_button.clicked.connect(self.start_processing)

        # 진행 상황 표시
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        # 왼쪽 레이아웃에 위젯 추가
        left_layout.addWidget(self.rabbit_label)
        left_layout.addWidget(self.skd_button)
        left_layout.addWidget(self.skd_label)
        left_layout.addWidget(self.dna_bs_button)
        left_layout.addWidget(self.dna_bs_label)
        left_layout.addWidget(self.dna_is_button)
        left_layout.addWidget(self.dna_is_label)
        left_layout.addWidget(self.output_dir_button)
        left_layout.addWidget(self.output_dir_label)
        left_layout.addWidget(self.start_button)
        left_layout.addWidget(self.progress_bar)

        # --- 오른쪽 레이아웃 ---
        # 로그 출력 뷰 (크게)
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        font = QFont()
        font.setPointSize(12)
        self.log_view.setFont(font)

        # 오른쪽 레이아웃에 위젯 추가
        right_layout.addWidget(self.log_view)

        # 메인 레이아웃에 왼쪽/오른쪽 레이아웃 추가
        main_layout.addLayout(left_layout)
        main_layout.addLayout(right_layout)
        self.setLayout(main_layout)

        # 파일 경로 변수 초기화
        self.skd_filepath = None
        self.dna_bs_filepath = None
        self.dna_is_filepath = None
        self.output_dir = None

    def select_skd_file(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Data 파일 선택", "", "Excel 파일 (*.xlsx)")
        if filepath:
            self.skd_filepath = filepath
            self.skd_label.setText(f"Data 파일: {os.path.basename(filepath)}")

    def select_dna_bs_file(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "DNA BS 파일 선택", "", "Excel 파일 (*.xlsx)")
        if filepath:
            self.dna_bs_filepath = filepath
            self.dna_bs_label.setText(f"DNA BS 파일: {os.path.basename(filepath)}")

    def select_dna_is_file(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "DNA IS 파일 선택", "", "Excel 파일 (*.xlsx)")
        if filepath:
            self.dna_is_filepath = filepath
            self.dna_is_label.setText(f"DNA IS 파일: {os.path.basename(filepath)}")

    def select_output_dir(self):
        dirpath = QFileDialog.getExistingDirectory(self, "결과 파일 저장 위치 선택")
        if dirpath:
            self.output_dir = dirpath
            self.output_dir_label.setText(f"저장 위치: {dirpath}")

    def start_processing(self):
        if not all([self.skd_filepath, self.dna_bs_filepath, self.dna_is_filepath, self.output_dir]):
            QMessageBox.warning(self, "파일 선택", "모든 파일 및 저장 위치를 선택해주세요.")
            return

        self.progress_bar.setVisible(True)
        self.start_button.setEnabled(False)

        self.worker_thread = WorkerThread(self.skd_filepath, self.dna_bs_filepath, self.dna_is_filepath, self.output_dir)
        self.worker_thread.progress_updated.connect(self.update_progress)
        self.worker_thread.log_updated.connect(self.update_log)
        self.worker_thread.finished.connect(self.processing_finished)
        self.worker_thread.start()

    def update_progress(self, progress):
        self.progress_bar.setValue(progress)

    def update_log(self, log_message):
        self.log_view.append(log_message)

    def processing_finished(self):
        self.progress_bar.setVisible(False)
        self.start_button.setEnabled(True)
        QMessageBox.information(self, "작업 완료", "데이터 처리가 완료되었습니다!")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())