import sys
import re
import json
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout, 
                             QWidget, QLabel, QTextEdit, QFileDialog, QMessageBox, 
                             QListWidget, QTabWidget, QDialog, QCheckBox, QDialogButtonBox,
                             QMenu, QAction, QInputDialog, QLineEdit)
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import Qt, QThread, pyqtSignal
import numpy as np
from paddleocr import PaddleOCR
import fitz
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime

class OCRThread(QThread):
    ocr_complete = pyqtSignal(list)

    def __init__(self, ocr, image):
        super().__init__()
        self.ocr = ocr
        self.image = image

    def run(self):
        result = self.ocr.ocr(self.image, cls=True)
        self.ocr_complete.emit(result[0] if result else [])

class ProgressPopup(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent, Qt.WindowStaysOnTopHint)
        self.setWindowTitle("처리 중")
        self.setFixedSize(200, 100)
        layout = QVBoxLayout(self)
        self.label = QLabel("OCR 작업 진행 중...")
        layout.addWidget(self.label)
        self.setModal(True)

class EditableListWidget(QListWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

    def show_context_menu(self, pos):
        global_pos = self.mapToGlobal(pos)
        menu = QMenu()
        edit_action = menu.addAction("Edit")
        copy_action = menu.addAction("Copy")
        
        action = menu.exec_(global_pos)
        
        if action == edit_action:
            self.edit_item()
        elif action == copy_action:
            self.copy_item()

    def edit_item(self):
        current_item = self.currentItem()
        if current_item:
            text = current_item.text()
            if ':' in text:
                key, value = text.split(':', 1)
                new_value, ok = QInputDialog.getText(self, "Edit Item", "Enter new value:", QLineEdit.Normal, value.strip())
                if ok and new_value:
                    current_item.setText(f"{key}: {new_value}")
                    # Update the accumulated_data dictionary
                    year = self.get_current_year()
                    if year:
                        self.parent().parent().parent().accumulated_data[year][key.strip()] = new_value.strip()

    def copy_item(self):
        current_item = self.currentItem()
        if current_item:
            QApplication.clipboard().setText(current_item.text())

    def get_current_year(self):
        for i in range(self.currentRow(), -1, -1):
            item_text = self.item(i).text()
            if item_text.startswith('---') and item_text.endswith('---'):
                return item_text.strip('- ')
        return None

class FinancialStatementApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("노가다 헬퍼")
        self.setGeometry(100, 100, 1500, 800)

        self.ocr = PaddleOCR(use_angle_cls=True, lang="ch")
        
        self.file_path = None
        self.current_page = 0
        self.total_pages = 0
        self.pages = []
        self.ocr_results = []
        self.mapped_data = {}
        self.accumulated_data = {}
        self.statement_type = None
        self.excel_file_path = None
        self.mapping = {
            "货币资金": "현금 및 현금성 자산",
            "货币现金": "현금 및 현금성 자산",
            "以公允价值计量变动投资的金融资金": "현금 및 현금성 자산",
            "应收票据": "받을어음",
            "应收账款": "외상매출금",
            "应收款项融资": "받을어음",
            "其他应收款": "기타미수금",
            "应收利息": "미수이자",
            "待摊费用": "선급금",
            "预付款项": "선급금",
            "应收股利": "미수배당금",
            "一年内到期的长债权": "유동성비유동자산",
            "其他流动资产": "기타유동자산",
            "存货": "재고자산",
            "可供出货金融资金": "매도가능증권",
            "长期股权投资": "장기투자주식",
            "房地产投资": "투자부동산",
            "长期债券投资": "기타 투자자산",
            "其他权益工具投资": "기타 투자자산",
            "在建工程": "건설중인자산",
            "固定资产净额": "고정자산 정리",
            "固定资产": "고정자산 정리",
            "固定资产账面价值": "고정자산 정리",
            "其他固定资产": "기타유형자산",
            "使用权资产": "기타의 무형자산",
            "商誉": "영업권",
            "开发支出": "개발비",
            "无形资产": "기타의 무형자산",
            "长期待摊费用": "장기선급비용",
            "递延所得税资产": "이연법인세자산",
            "其他非流动资产": "기타 비유동자산",
            "短期借款": "단기차입금",
            "应付账款": "외상매입금",
            "应付票据": "지급어음",
            "预收账款": "선수금",
            "合同负债": "선수금",
            "预提费用": "부채성충당금",
            "预计负债": "부채성충당금",
            "应付职工薪酬": "미지급인건비",
            "应交税费": "미지급세금",
            "应付利息": "미지급이자",
            "应付股利": "미지급배당금",
            "应付利润": "미지급배당금",
            "其他应付款": "기타미지급금",
            "其他应交款": "기타미지급금",
            "一年内到期的长期负债": "유동성비유동부채",
            "交易性金融负债": "단기매매부채",
            "其他流动负债金额": "기타유동부채금액",
            "长期借款": "장기차입금",
            "应付债券": "사채",
            "长期应付款": "장기미지급금",
            "递延所得税负债": "이연법인세부채",
            "租赁负债": "기타비유동부채금액",
            "递延收益": "기타비유동부채금액",
            "实收资金": "자본금",
            "股本": "자본금",
            "资本公积": "자본잉여금",
            "其他权益工具": "자본잉여금",
            "专项储备": "자본잉여금",
            "盈余公积": "이익잉여금",
            "未分配盈利": "미처분이익잉여금"
        }

        self.progress_popup = None
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout()

        # Left side (File selection, Image preview, and OCR control)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)

        self.file_btn = QPushButton("PDF/Image 선택")
        self.file_btn.clicked.connect(self.select_file)
        left_layout.addWidget(self.file_btn)

        nav_layout = QHBoxLayout()
        self.prev_btn = QPushButton("이전")
        self.prev_btn.clicked.connect(self.prev_page)
        self.next_btn = QPushButton("다음")
        self.next_btn.clicked.connect(self.next_page)
        self.page_label = QLabel("Page: 0/0")
        nav_layout.addWidget(self.prev_btn)
        nav_layout.addWidget(self.page_label)
        nav_layout.addWidget(self.next_btn)
        left_layout.addLayout(nav_layout)

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        left_layout.addWidget(self.image_label)

        self.ocr_btn = QPushButton("OCR 실행!")
        self.ocr_btn.clicked.connect(self.perform_ocr)
        left_layout.addWidget(self.ocr_btn)

        # Right side (Tabbed widget for OCR results, Mapped data, Accumulated data, and Excel processing)
        right_widget = QTabWidget()

        # OCR Results tab
        ocr_tab = QWidget()
        ocr_layout = QVBoxLayout(ocr_tab)
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        ocr_layout.addWidget(self.results_text)
        right_widget.addTab(ocr_tab, "OCR 결과")

        # Mapped Data tab
        mapped_tab = QWidget()
        mapped_layout = QVBoxLayout(mapped_tab)
        self.mapped_data_list = QListWidget()
        mapped_layout.addWidget(self.mapped_data_list)

        self.accumulate_btn = QPushButton("임시저장 데이터")
        self.accumulate_btn.clicked.connect(self.accumulate_data)
        mapped_layout.addWidget(self.accumulate_btn)

        right_widget.addTab(mapped_tab, "맵핑된 데이터")

        # Accumulated Data tab
        accumulated_tab = QWidget()
        accumulated_layout = QVBoxLayout(accumulated_tab)
        self.accumulated_data_list = EditableListWidget()
        accumulated_layout.addWidget(self.accumulated_data_list)
        right_widget.addTab(accumulated_tab, "임시저장 데이터")

        # Excel Processing tab
        excel_tab = QWidget()
        excel_layout = QVBoxLayout(excel_tab)
        
        self.excel_file_btn = QPushButton("Excel File 선택")
        self.excel_file_btn.clicked.connect(self.select_excel_file)
        excel_layout.addWidget(self.excel_file_btn)

        self.process_excel_btn = QPushButton("Excel 수정")
        self.process_excel_btn.clicked.connect(self.process_excel)
        excel_layout.addWidget(self.process_excel_btn)

        self.excel_result_text = QTextEdit()
        self.excel_result_text.setReadOnly(True)
        excel_layout.addWidget(self.excel_result_text)

        right_widget.addTab(excel_tab, "Excel")

        # Add widgets to main layout
        main_layout.addWidget(left_widget, 1)
        main_layout.addWidget(right_widget, 1)

        central_widget.setLayout(main_layout)

    def select_file(self):
        file_dialog = QFileDialog()
        self.file_path, _ = file_dialog.getOpenFileName(self, "Select PDF/Image", "", "PDF Files (*.pdf);;Image Files (*.png *.jpg *.jpeg)")
        if self.file_path:
            if self.file_path.lower().endswith('.pdf'):
                self.load_pdf()
            else:
                self.load_image()

    def load_pdf(self):
        try:
            pdf_document = fitz.open(self.file_path)
            self.total_pages = len(pdf_document)
            self.pages = [pdf_document[i].get_pixmap() for i in range(self.total_pages)]
            self.current_page = 0
            self.update_page_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load PDF: {str(e)}")

    def load_image(self):
        try:
            image = Image.open(self.file_path)
            self.pages = [image]
            self.total_pages = 1
            self.current_page = 0
            self.update_page_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load image: {str(e)}")

    def update_page_display(self):
        if self.pages:
            if isinstance(self.pages[self.current_page], fitz.Pixmap):
                img = Image.frombytes("RGB", [self.pages[self.current_page].width, self.pages[self.current_page].height], self.pages[self.current_page].samples)
            else:
                img = self.pages[self.current_page]
            self.display_image(img)
            self.page_label.setText(f"Page: {self.current_page + 1}/{self.total_pages}")

    def display_image(self, img):
        img = img.convert("RGB")
        width, height = img.size
        aspect_ratio = width / height
        new_width = 600
        new_height = int(new_width / aspect_ratio)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        qimage = QImage(img.tobytes(), img.width, img.height, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qimage)
        self.image_label.setPixmap(pixmap)

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_page_display()

    def next_page(self):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.update_page_display()

    def perform_ocr(self):
        if not self.pages:
            QMessageBox.warning(self, "No File", "Please select a PDF or image file first.")
            return

        if isinstance(self.pages[self.current_page], fitz.Pixmap):
            img = Image.frombytes("RGB", [self.pages[self.current_page].width, self.pages[self.current_page].height], self.pages[self.current_page].samples)
        else:
            img = self.pages[self.current_page]
        
        img_np = np.array(img)

        self.ocr_thread = OCRThread(self.ocr, img_np)
        self.ocr_thread.ocr_complete.connect(self.process_ocr_results)
        self.ocr_thread.start()

        self.progress_popup = ProgressPopup(self)
        self.progress_popup.show()

    def process_ocr_results(self, results):
        self.ocr_results = results
        self.display_ocr_results()
        self.auto_match_items_values()
        if self.progress_popup:
            self.progress_popup.close()
            self.progress_popup = None

    def display_ocr_results(self):
        self.results_text.clear()
        for line in self.ocr_results:
            text = line[1][0]
            confidence = line[1][1]
            self.results_text.append(f"{text} (Confidence: {confidence:.2f})")

    def auto_match_items_values(self):
        self.mapped_data = {}
        current_item = None
        years = []
        
        # Determine statement type and extract years
        for line in self.ocr_results:
            text = line[1][0]
            if "资产负债表" in text:
                self.statement_type = "资产负债表"
                break
            elif "利润表" in text:
                self.statement_type = "利润表"
                break

        # Extract years
        for line in self.ocr_results:
            text = line[1][0]
            year_match = re.search(r'\b(\d{4})年度?\b', text)
            if year_match:
                year = year_match.group(1)
                if year not in years:
                    years.append(year)
                    self.mapped_data[year] = {}

        if not years:
            QMessageBox.warning(self, "Error", "No years detected in the document.")
            return

        years.sort(reverse=True)

        for line in self.ocr_results:
            text = line[1][0]
            
            # Identify item name
            if re.match(r'^[\u4e00-\u9fff:：()（）\s]+$', text) and len(text) > 1:
                current_item = text.strip()
            
            # Identify value
            elif re.match(r'^[-]?\(?[\d,]+\.?\d*\)?$', text):
                if current_item:
                    value = text.strip('()')
                    value = value.strip(',')
                    
                    # Assign value to years
                    for year in years:
                        if current_item not in self.mapped_data[year]:
                            self.mapped_data[year][current_item] = value
                            break
        
        self.update_mapped_data_view()

    def update_mapped_data_view(self):
        self.mapped_data_list.clear()
        for year in sorted(self.mapped_data.keys(), reverse=True):
            self.mapped_data_list.addItem(f"--- {year} ---")
            for item_name, value in self.mapped_data[year].items():
                self.mapped_data_list.addItem(f"{item_name}: {value}")
            self.mapped_data_list.addItem("")  # Add empty line between years

    def accumulate_data(self):
        if not self.mapped_data:
            QMessageBox.warning(self, "No Data", "No data to accumulate. Please perform OCR first.")
            return

        # Create a dialog for year selection
        year_dialog = QDialog(self)
        year_dialog.setWindowTitle("Select Years to Accumulate")
        layout = QVBoxLayout(year_dialog)

        year_checkboxes = {}
        for year in sorted(self.mapped_data.keys(), reverse=True):
            checkbox = QCheckBox(str(year))
            checkbox.setChecked(True)  # Default to checked
            year_checkboxes[year] = checkbox
            layout.addWidget(checkbox)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(year_dialog.accept)
        button_box.rejected.connect(year_dialog.reject)
        layout.addWidget(button_box)

        if year_dialog.exec_() == QDialog.Accepted:
            selected_years = [year for year, checkbox in year_checkboxes.items() if checkbox.isChecked()]
            
            if not selected_years:
                QMessageBox.warning(self, "No Years Selected", "Please select at least one year to accumulate data.")
                return

            error_messages = []

            for year in selected_years:
                if year not in self.accumulated_data:
                    self.accumulated_data[year] = {}
                for item, value in self.mapped_data[year].items():
                    try:
                        if item in self.accumulated_data[year]:
                            self.accumulated_data[year][item] = self.safe_float(self.accumulated_data[year][item]) + self.safe_float(value)
                        else:
                            self.accumulated_data[year][item] = self.safe_float(value)
                    except Exception as e:
                        error_messages.append(f"Error processing {item} for year {year}: {str(e)}")

            if error_messages:
                error_text = "\n".join(error_messages)
                QMessageBox.warning(self, "Accumulation Warnings", f"Some items could not be processed:\n\n{error_text}")
            else:
                QMessageBox.information(self, "Accumulation Complete", "Data has been accumulated successfully.")

            self.update_accumulated_data_view()

    def update_accumulated_data_view(self):
        self.accumulated_data_list.clear()
        for year in sorted(self.accumulated_data.keys(), reverse=True):
            self.accumulated_data_list.addItem(f"--- {year} ---")
            for item_name, value in self.accumulated_data[year].items():
                self.accumulated_data_list.addItem(f"{item_name}: {value}")
            self.accumulated_data_list.addItem("")  # Add empty line between years

    def select_excel_file(self):
        file_dialog = QFileDialog()
        self.excel_file_path, _ = file_dialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if self.excel_file_path:
            QMessageBox.information(self, "File Selected", f"Excel file selected: {self.excel_file_path}")

    def process_excel(self):
        if not self.excel_file_path:
            QMessageBox.warning(self, "No Excel File", "Please select an Excel file first.")
            return

        if not self.accumulated_data:
            QMessageBox.warning(self, "No Data", "No accumulated data to process. Please perform OCR and accumulate data first.")
            return

        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sheet = workbook.active

            current_year = datetime.datetime.now().year
            cleaned_mapping = {self.clean_key_string(v): k for k, v in self.mapping.items()}

            for year in sorted(self.accumulated_data.keys(), reverse=True):
                column_letter = get_column_letter(int(year) - current_year + 7)

                for row in range(2, sheet.max_row + 1):
                    korean_key = sheet.cell(row=row, column=1).value
                    
                    if korean_key:
                        cell = sheet.cell(row=row, column=column_index_from_string(column_letter))
                        cumulative_value = 0

                        matched = False
                        for chinese_key, mapped_korean_key in self.mapping.items():
                            if self.clean_key_string(mapped_korean_key) == self.clean_key_string(korean_key):
                                if chinese_key in self.accumulated_data[year]:
                                    value = self.safe_float(self.accumulated_data[year][chinese_key])
                                    cumulative_value += value
                                    matched = True
                                    self.excel_result_text.append(f"매칭: {korean_key} -> {chinese_key}, 값: {value}")

                        if matched:
                            cell.value = cumulative_value  # 기존 값을 더하지 않고 새 값으로 대체
                            self.excel_result_text.append(f"최종 값: {korean_key} = {cell.value}")
                        else:
                            self.excel_result_text.append(f"매칭 실패: {korean_key}")

            output_excel_file_path = self.excel_file_path.replace('.xlsx', '_output.xlsx')
            workbook.save(output_excel_file_path)
            self.excel_result_text.append(f"처리 완료. 결과가 {output_excel_file_path}에 저장되었습니다.")
            QMessageBox.information(self, "Processing Complete", f"Excel file has been processed and saved as {output_excel_file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while processing the Excel file: {str(e)}")

    def clean_key_string(self, key):
        if key is None:
            return ""
        cleaned = re.sub(r'^[\s\d().]+|[\s\d().]+$', '', str(key))
        cleaned = re.sub(r'\s+', ' ', cleaned)
        return cleaned.strip()

    def safe_float(self, value):
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(',', ''))
        except (ValueError, AttributeError):
            return 0.0

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Message',
            "Are you sure you want to quit?", QMessageBox.Yes |
            QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FinancialStatementApp()
    window.show()
    sys.exit(app.exec_())  